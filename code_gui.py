import os
import json
import pymupdf
import fitz  # This is the PyMuPDF library
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Border, Fill, Protection, Alignment, Side
import shutil
import datetime

# --- GUI Imports ---
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading

# -------------------------------------------------------------------
# --- ALL YOUR ORIGINAL HELPER FUNCTIONS (UNCHANGED) ---
# -------------------------------------------------------------------

def get_resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def load_config(config_path):
    """Loads the configuration file."""
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: Configuration file '{config_path}' not found.")
        print("Please create a 'config.json' file in the same directory.")
        return None
    except json.JSONDecodeError:
        print(f"Error: Could not parse '{config_path}'. Make sure it is valid JSON.")
        return None
    # We remove the setup_directories function as the user will provide the path.

def extract_data_from_pdf(pdf_path, fields):
    """
    Extracts data from a single PDF based on the fields defined in the config.
    """
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print(f"Error opening {pdf_path}: {e}")
        return None

    extracted_data = {}
    for field in fields:
        try:
            page_num = field['page']
            rect_coords = field['rect']
            
            if page_num >= doc.page_count:
                print(f"Warning: Page {page_num} out of range for {pdf_path}. Skipping field '{field['name']}'.")
                extracted_data[field['name']] = ""
                continue
                
            page = doc.load_page(page_num)
            rect = fitz.Rect(rect_coords[0], rect_coords[1], rect_coords[2], rect_coords[3])
            text = page.get_text("text", clip=rect).strip()
            text = text.replace('\n', ' ').replace('\r', ' ')
            extracted_data[field['name']] = text
        except Exception as e:
            print(f"Error processing field '{field['name']}' in {pdf_path}: {e}")
            extracted_data[field['name']] = "ERROR"

    doc.close()
    return extracted_data

# --- Excel Helper Functions (UNCHANGED) ---

def copy_sheet_properties(source_sheet, target_sheet):
    """
    Manually copies cell values, styles, dimensions, and merged cells
    from source_sheet to target_sheet (which must be in different workbooks).
    """
    print(f"Copying properties from '{source_sheet.title}' to '{target_sheet.title}'...")
    
    # 1. Copy cell values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, vertAlign=cell.font.vertAlign, underline=cell.font.underline, strike=cell.font.strike, color=cell.font.color)
                    new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom, diagonal=cell.border.diagonal, diagonal_direction=cell.border.diagonal_direction, outline=cell.border.outline, start=cell.border.start, end=cell.border.end)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = Protection(locked=cell.protection.locked, hidden=cell.protection.hidden)
                    new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, text_rotation=cell.alignment.text_rotation, wrap_text=cell.alignment.wrap_text, shrink_to_fit=cell.alignment.shrink_to_fit, indent=cell.alignment.indent)
                except Exception as e:
                    print(f"Warning: Could not copy style for cell {cell.coordinate}. Error: {e}")

    # 2. Copy column dimensions
    for col_letter, dim in source_sheet.column_dimensions.items():
        if dim.width: target_sheet.column_dimensions[col_letter].width = dim.width
        target_sheet.column_dimensions[col_letter].hidden = dim.hidden

    # 3. Copy row dimensions
    for row_index, dim in source_sheet.row_dimensions.items():
        if dim.height: target_sheet.row_dimensions[row_index].height = dim.height
        target_sheet.row_dimensions[row_index].hidden = dim.hidden

    # 4. Copy merged cells
    for merge_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merge_range))
    
    print(f"Finished copying properties to '{target_sheet.title}'.")

def get_header_map(sheet, header_row=4):
    """
    Reads the header row and returns a dictionary mapping
    header names to column indices. e.g. {'FECHA': 1, 'BATEA': 3}
    """
    header_map = {}
    for cell in sheet[header_row]:
        if cell.value:
            header_name = str(cell.value).strip()
            header_map[header_name] = cell.column
    return header_map

def write_data_to_sheet(sheet, header_map, data_row, target_row):
    """
    Writes a single PDF's data (data_row) to the specified sheet
    at the target_row, using the header_map to find correct columns.
    """
    sheet.insert_rows(target_row)
    black_font = Font(color="000000")
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for col_idx in range(1, 15): 
        cell = sheet.cell(row=target_row, column=col_idx)
        cell.border = thin_border
    
    field_to_header_map = {
        "FECHA": "FECHA", "BATEA": "BATEA", "EMPRESA": "EMPRESA",
        "NIF EMPRESA": "NIF EMPRESA", "KG BRUTOS": "KG BRUTOS",
        "DESCUENTO": "DESCUENTO", "KG NETOS": "KG NETOS"
    }

    for field_name, header_name in field_to_header_map.items():
        if header_name in header_map:
            col_idx = header_map[header_name]
            value_to_write = data_row.get(field_name, "")
            
            try:
                numeric_value = float(value_to_write)
                cell = sheet.cell(row=target_row, column=col_idx, value=numeric_value)
                cell.font = black_font
            except (ValueError, TypeError):
                cell = sheet.cell(row=target_row, column=col_idx, value=data_row.get(field_name, ""))
                cell.font = black_font
        else:
            print(f"Warning: Header '{header_name}' not found in sheet '{sheet.title}'. Skipping data.")

def get_or_create_sheet(workbook, batea_name, template_sheet_name="TEMPLATE"):
    """
    Tries to find a sheet with batea_name. If not found,
    copies the internal 'TEMPLATE' sheet and renames it.
    """
    if batea_name in workbook.sheetnames:
        return workbook[batea_name]
    else:
        print(f"Sheet '{batea_name}' not found. Creating it from '{template_sheet_name}'...")
        
        if template_sheet_name not in workbook.sheetnames:
            print(f"CRITICAL ERROR: Template sheet '{template_sheet_name}' not found in the workbook.")
            return None 
            
        template_sheet = workbook[template_sheet_name]
        
        if not hasattr(template_sheet, 'parent') or template_sheet.parent != workbook:
            print(f"CRITICAL ERROR: The template sheet '{template_sheet_name}' is not a valid worksheet.")
            return None
            
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = batea_name
        return new_sheet

# -------------------------------------------------------------------
# --- NEW GUI APPLICATION CLASS ---
# -------------------------------------------------------------------

class TextRedirector:
    """A helper class to redirect stdout/stderr to a Tkinter Text widget."""
    def __init__(self, widget):
        self.widget = widget

    def write(self, s):
        # Ensure GUI updates happen on the main thread
        self.widget.after(0, self.append_text, s)

    def append_text(self, s):
        self.widget.config(state="normal")
        self.widget.insert(tk.END, s)
        self.widget.see(tk.END)
        self.widget.config(state="disabled")

    def flush(self):
        pass  # Required for file-like object


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF to Excel Extractor")
        self.root.geometry("700x500") # Width x Height

        self.root.resizable(False, False)

        # --- Configure style ---
        self.style = ttk.Style()
        self.style.theme_use('clam') # 'clam', 'alt', 'default', 'classic'

        # --- Main frame ---
        self.main_frame = ttk.Frame(root, padding="10 10 10 10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # --- 1. PDF Folder Selection ---
        self.pdf_frame = ttk.LabelFrame(self.main_frame, text="1. Select PDF Folder", padding="10")
        self.pdf_frame.pack(fill="x", expand=False, pady=5)
        
        self.pdf_folder_path = tk.StringVar()
        self.pdf_entry = ttk.Entry(self.pdf_frame, textvariable=self.pdf_folder_path, width=60)
        self.pdf_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 5))
        
        self.pdf_button = ttk.Button(self.pdf_frame, text="Browse", width=10, command=self.browse_pdf_folder)
        self.pdf_button.pack(side=tk.LEFT)

        # --- 2. Excel File Selection ---
        self.excel_frame = ttk.LabelFrame(self.main_frame, text="2. Select/Create Excel Output File", padding="10")
        self.excel_frame.pack(fill="x", expand=False, pady=5)
        
        self.excel_file_path = tk.StringVar()
        self.excel_entry = ttk.Entry(self.excel_frame, textvariable=self.excel_file_path, width=60)
        self.excel_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=(0, 5))
        
        self.excel_button = ttk.Button(self.excel_frame, text="Browse", width=10, command=self.browse_excel_file)
        self.excel_button.pack(side=tk.LEFT)

        # --- 3. Start Button ---
        self.start_button = ttk.Button(self.main_frame, text="Start Processing", command=self.start_processing_thread)
        self.start_button.pack(pady=10, fill="x")

        # --- 4. Log/Status Output ---
        self.log_frame = ttk.LabelFrame(self.main_frame, text="Log", padding="10")
        self.log_frame.pack(fill="both", expand=True, pady=5)
        
        self.status_text = scrolledtext.ScrolledText(self.log_frame, state="disabled", height=15, wrap=tk.WORD)
        self.status_text.pack(fill="both", expand=True)
        
        # --- Redirect stdout/stderr to the text widget ---
        self.redirector = TextRedirector(self.status_text)
        sys.stdout = self.redirector
        sys.stderr = self.redirector
        
        print("Ready. Please select your folders and files.")

    def browse_pdf_folder(self):
        """Opens a dialog to select the PDF input folder."""
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.pdf_folder_path.set(folder_selected)
            print(f"PDF folder set to: {folder_selected}")

    def browse_excel_file(self):
        """
        Opens a dialog to select an existing Excel file or
        name a new one.
        """
        file_selected = filedialog.asksaveasfilename(
            title="Select or name your output Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )
        if file_selected:
            self.excel_file_path.set(file_selected)
            print(f"Excel file set to: {file_selected}")

    def start_processing_thread(self):
        """Validates input and starts the main logic in a new thread."""
        pdf_path = self.pdf_folder_path.get()
        excel_path = self.excel_file_path.get()

        # --- Validation ---
        if not pdf_path or not excel_path:
            messagebox.showerror("Error", "Please select both a PDF folder and an Excel output file.")
            return
            
        if not os.path.isdir(pdf_path):
            messagebox.showerror("Error", f"The selected PDF folder does not exist:\n{pdf_path}")
            return

        # Disable button to prevent double-clicks
        self.start_button.config(state="disabled", text="Processing...")
        
        # Clear the log
        self.status_text.config(state="normal")
        self.status_text.delete('1.0', tk.END)
        self.status_text.config(state="disabled")

        # Run the main logic in a separate thread
        self.processing_thread = threading.Thread(
            target=self.run_main_logic,
            args=(pdf_path, excel_path),
            daemon=True
        )
        self.processing_thread.start()

    def run_main_logic(self, input_folder, output_filename):
        """
        This is your original 'main()' function, refactored to run as a
        method and provide feedback to the GUI.
        """
        try:
            print("Starting PDF processing...")
            
            config_path = get_resource_path('config.json')
            config = load_config(config_path)
            
            if config is None:
                messagebox.showerror("Config Error", "Failed to load 'config.json'. Check the log for details.")
                return # Exit thread

            all_data = []
            
            pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
            
            if not pdf_files:
                print(f"No PDF files found in '{input_folder}'.")
                messagebox.showwarning("No Files", f"No PDF files were found in '{input_folder}'.")
                return # Exit thread

            print(f"Found {len(pdf_files)} PDF(s) to process...")

            for filename in pdf_files:
                pdf_path = os.path.join(input_folder, filename)
                print(f"Processing '{filename}'...")
                data = extract_data_from_pdf(pdf_path, config['extraction_fields'])
                
                if data:
                    data["FECHA"] = data["FECHA"].split(" ")[0]
                    data['Source File'] = filename
                    all_data.append(data)

            if not all_data:
                print("No data was successfully extracted from any PDF.")
                messagebox.showinfo("Finished", "Processing complete, but no data was extracted.")
                return # Exit thread

            print("Data extraction complete. Generating Excel file...")
            
            template_path = get_resource_path('PLANTILLA.xlsx')
            if not os.path.exists(template_path):
                print(f"Error: Template file 'PLANTILLA.xlsx' not found.")
                messagebox.showerror("Template Error", "Template file 'PLANTILLA.xlsx' not found. Please make sure it is in the same directory as the application.")
                return
            
            # --- Logic for creating or updating file ---
            
            if os.path.exists(output_filename):
                # --- MODE 2: UPDATE EXISTING FILE ---
                print(f"File '{output_filename}' exists. Will update it.")
                
                try:
                    backup_dir = ".old"
                    os.makedirs(backup_dir, exist_ok=True)
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
                    base_name, extension = os.path.splitext(output_filename)
                    backup_filename = f"{base_name}_{timestamp}{extension}"
                    backup_path = os.path.join(backup_dir, backup_filename)
                    shutil.copy(output_filename, backup_path)
                    print(f"Created backup of '{output_filename}' at '{backup_path}'")
                except Exception as e:
                    print(f"Warning: Could not create backup for '{output_filename}'. Error: {e}")
                
                try:
                    workbook = load_workbook(output_filename)
                except (InvalidFileException, FileNotFoundError):
                    print(f"Error: Could not open '{output_filename}'. It might be corrupt or not a valid .xlsx file.")
                    messagebox.showerror("File Error", f"Could not open '{output_filename}'. It might be corrupt or not a valid .xlsx file.")
                    return
                except Exception as e:
                    print(f"Error loading workbook: {e}")
                    messagebox.showerror("File Error", f"Error loading workbook: {e}")
                    return
                
                template_sheet_name = "TEMPLATE"
                if template_sheet_name not in workbook.sheetnames:
                    print(f"Warning: '{template_sheet_name}' sheet not found in '{output_filename}'.")
                    print("Attempting to create it from 'PLANTILLA.xlsx'...")
                    
                    try:
                        if not os.path.exists(template_path):
                             print(f"CRITICAL ERROR: 'PLANTILLA.xlsx' not found at {template_path}. Cannot create 'TEMPLATE' sheet.")
                             messagebox.showerror("Template Error", f"CRITICAL ERROR: 'PLANTILLA.xlsx' not found at {template_path}. Cannot create 'TEMPLATE' sheet.")
                             return

                        template_wb = load_workbook(template_path)
                        source_sheet = template_wb.active
                        target_sheet = workbook.create_sheet(title=template_sheet_name)
                        copy_sheet_properties(source_sheet, target_sheet)
                        template_wb.close()
                        print(f"Successfully created 'TEMPLATE' sheet in '{output_filename}'.")
                    
                    except Exception as e:
                        print(f"CRITICAL ERROR: Failed to create 'TEMPLATE' sheet from 'PLANTILLA.xlsx'.")
                        print(f"Error details: {e}")
                        messagebox.showerror("Template Error", f"CRITICAL ERROR: Failed to create 'TEMPLATE' sheet from 'PLANTILLA.xlsx'.\n{e}")
                        return
                
                for data_row in all_data:
                    batea = data_row.get("BATEA", "").strip()
                    if not batea:
                        print(f"Warning: PDF '{data_row['Source File']}' has no BATEA. Skipping.")
                        continue
                    
                    sheet = get_or_create_sheet(workbook, batea, template_sheet_name)
                    
                    if sheet is None:
                        print(f"Skipping PDF '{data_row['Source File']}' due to missing template sheet.")
                        continue
                        
                    header_map = get_header_map(sheet, header_row=4)
                    target_row = 5            
                    print(f"Writing data from '{data_row['Source File']}' to sheet '{batea}', row {target_row}...")
                    write_data_to_sheet(sheet, header_map, data_row, target_row)

            else:
                # --- MODE 1: CREATE NEW FILE ---
                print(f"Creating new file '{output_filename}' from template...")
                
                data_by_batea = {}
                for row in all_data:
                    batea = row.get('BATEA', 'UNKNOWN_BATEA').strip()
                    if not batea:
                        batea = 'UNKNOWN_BATEA'
                    if batea not in data_by_batea:
                        data_by_batea[batea] = []
                    data_by_batea[batea].append(row)
                
                shutil.copy(template_path, output_filename)
                
                try:
                    workbook = load_workbook(output_filename)
                    template_sheet = workbook.active
                    template_sheet_name = "TEMPLATE"
                    template_sheet.title = template_sheet_name
                except (InvalidFileException, FileNotFoundError):
                    print(f"Error: Could not open the new file '{output_filename}'.")
                    messagebox.showerror("File Error", f"Error: Could not open the new file '{output_filename}'.")
                    return
                except Exception as e:
                    print(f"Error loading new workbook: {e}")
                    messagebox.showerror("File Error", f"Error loading new workbook: {e}")
                    return
                
                for batea_name, batea_data in data_by_batea.items():
                    print(f"Creating sheet '{batea_name}'...")
                    sheet = get_or_create_sheet(workbook, batea_name, template_sheet_name)
                    
                    if sheet is None:
                        print(f"Skipping Batea '{batea_name}' due to missing template sheet.")
                        continue
                    
                    header_map = get_header_map(sheet, header_row=4)
                    
                    target_row = 5
                    for data_row in batea_data:
                        print(f"Writing data from '{data_row['Source File']}' to sheet '{batea_name}', row {target_row}...")
                        write_data_to_sheet(sheet, header_map, data_row, target_row)

            # --- 4. Save and Exit ---
            try:
                if "TEMPLATE" in workbook.sheetnames:
                    print("Removing internal 'TEMPLATE' sheet...")
                    del workbook["TEMPLATE"]

                workbook.save(output_filename)
                print(f"\nSuccess! Data saved to '{output_filename}'")
                messagebox.showinfo("Success", f"Processing complete!\nData saved to '{output_filename}'")
            except PermissionError:
                print(f"\nError: Could not save '{output_filename}'.")
                print("Please make sure the file is not open in Excel.")
                messagebox.showerror("Permission Error", f"Could not save '{output_filename}'.\nPlease make sure the file is not open in Excel.")
            except Exception as e:
                print(f"\nError saving Excel file: {e}")
                messagebox.showerror("Save Error", f"An error occurred while saving the Excel file: {e}")

            workbook.close()
            print("Processing finished.")

        except Exception as e:
            # Catch-all for any unexpected errors
            print(f"An unexpected error occurred: {e}")
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("Critical Error", f"An unexpected error occurred:\n{e}")
        
        finally:
            # Re-enable the button, whether it succeeded or failed
            # We must use 'after' to ensure this runs on the main GUI thread
            self.root.after(0, self.enable_button)

    def enable_button(self):
        """Helper to safely re-enable the start button from the main thread."""
        self.start_button.config(state="normal", text="Start Processing")


# -------------------------------------------------------------------
# --- MAIN EXECUTION ---
# -------------------------------------------------------------------

if __name__ == "__main__":
    # Restore original stdout/stderr on exit
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    
    root = tk.Tk()
    app = App(root)
    
    try:
        root.mainloop()
    finally:
        # Restore console output when GUI closes
        sys.stdout = original_stdout
        sys.stderr = original_stderr