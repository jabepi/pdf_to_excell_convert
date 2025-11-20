import os
import json
import pymupdf
import fitz  # This is the PyMuPDF library
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Border, Fill, Protection, Alignment, Side
import shutil
import datetime  # <-- NEW IMPORT

def get_resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def load_config(config_path):
    """Loads the configuration file."""
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: Configuration file '{config_path}' not found.")
        print("Please create a 'config.json' file in the same directory.")
        return None
    except json.JSONDecodeError:
        print(f"Error: Could not parse '{config_path}'. Make sure it is valid JSON.")
        return None

def setup_directories(input_dir):
    """Checks if the input directory exists, and creates it if not."""
    if not os.path.exists(input_dir):
        print(f"Input folder '{input_dir}' not found. Creating it.")
        os.makedirs(input_dir)
        print(f"Please add your PDF files to the '{input_dir}' folder and run again.")
        return False
    return True

def extract_data_from_pdf(pdf_path, fields):
    """
    Extracts data from a single PDF based on the fields defined in the config.
    Each 'field' specifies a page and a rectangle [x0, y0, x1, y1].
    """
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print(f"Error opening {pdf_path}: {e}")
        return None

    extracted_data = {}
    for field in fields:
        try:
            page_num = field['page']
            rect_coords = field['rect']
            
            if page_num >= doc.page_count:
                print(f"Warning: Page {page_num} out of range for {pdf_path}. Skipping field '{field['name']}'.")
                extracted_data[field['name']] = ""
                continue
                
            page = doc.load_page(page_num)

            
            # DEBUG
            # print(page.get_textpage().extractXML())
            # text = page.get_text_blocks()
            # print(f"Full text on page {page_num}:\n{text}\n--- End of page text ---\n")
            # exit(0)
            
            # Define the rectangle (x0, y0, x1, y1)
            rect = fitz.Rect(rect_coords[0], rect_coords[1], rect_coords[2], rect_coords[3])
 
            # Extract text from that rectangle
            text = page.get_text("text", clip=rect).strip()
            
            # Clean up text (replace newlines with spaces)
            text = text.replace('\n', ' ').replace('\r', ' ')
            
            # print(f"Extracted '{field['name']}': {text}")
            extracted_data[field['name']] = text
        except Exception as e:
            print(f"Error processing field '{field['name']}' in {pdf_path}: {e}")
            extracted_data[field['name']] = "ERROR"

    doc.close()
    return extracted_data

# --- New Excel Helper Functions ---

def copy_sheet_properties(source_sheet, target_sheet):
    """
    Manually copies cell values, styles, dimensions, and merged cells
    from source_sheet to target_sheet (which must be in different workbooks).
    """
    print(f"Copying properties from '{source_sheet.title}' to '{target_sheet.title}'...")
    
    # 1. Copy cell values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, vertAlign=cell.font.vertAlign, underline=cell.font.underline, strike=cell.font.strike, color=cell.font.color)
                    new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom, diagonal=cell.border.diagonal, diagonal_direction=cell.border.diagonal_direction, outline=cell.border.outline, start=cell.border.start, end=cell.border.end)
                    # new_cell.fill = Fill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = Protection(locked=cell.protection.locked, hidden=cell.protection.hidden)
                    new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, text_rotation=cell.alignment.text_rotation, wrap_text=cell.alignment.wrap_text, shrink_to_fit=cell.alignment.shrink_to_fit, indent=cell.alignment.indent)
                except Exception as e:
                    print(f"Warning: Could not copy style for cell {cell.coordinate}. Error: {e}")

    # 2. Copy column dimensions
    for col_letter, dim in source_sheet.column_dimensions.items():
        if dim.width: # Only copy if width is set
            target_sheet.column_dimensions[col_letter].width = dim.width
        target_sheet.column_dimensions[col_letter].hidden = dim.hidden

    # 3. Copy row dimensions
    for row_index, dim in source_sheet.row_dimensions.items():
        if dim.height: # Only copy if height is set
            target_sheet.row_dimensions[row_index].height = dim.height
        target_sheet.row_dimensions[row_index].hidden = dim.hidden

    # 4. Copy merged cells
    for merge_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merge_range))
    
    print(f"Finished copying properties to '{target_sheet.title}'.")

def get_header_map(sheet, header_row=4):
    """
    Reads the header row and returns a dictionary mapping
    header names to column indices. e.g. {'FECHA': 1, 'BATEA': 3}
    """
    header_map = {}
    for cell in sheet[header_row]:
        if cell.value:
            header_name = str(cell.value).strip()
            header_map[header_name] = cell.column
    return header_map

def write_data_to_sheet(sheet, header_map, data_row, target_row):
    """
    Writes a single PDF's data (data_row) to the specified sheet
    at the target_row, using the header_map to find correct columns.
    
    --- NEW: This function also applies a border to the range A5:N5. ---
    """
    sheet.insert_rows(target_row)
    # Define the standard black font
    black_font = Font(color="000000")

    # --- NEW: Define the border style ---
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )

    # --- NEW: Apply border to row 5 (A5:N5) ---
    # Loop from column 1 (A) to 14 (N)
    for col_idx in range(1, 15): 
        cell = sheet.cell(row=target_row, column=col_idx)
        cell.border = thin_border
    
    # --- End of new border logic ---


    # Maps config.json 'name' to Excel 'header name'
    # This allows config and Excel to have slightly different names if needed,
    # but here they are identical.
    field_to_header_map = {
        "FECHA": "FECHA",
        "BATEA": "BATEA",
        "EMPRESA": "EMPRESA",
        "NIF EMPRESA": "NIF EMPRESA",
        "KG BRUTOS": "KG BRUTOS",
        "DESCUENTO": "DESCUENTO",
        "KG NETOS": "KG NETOS"
    }

    for field_name, header_name in field_to_header_map.items():
        if header_name in header_map:
            col_idx = header_map[header_name]
            value_to_write = data_row.get(field_name, "")
            
            try:
                numeric_value = float(value_to_write)
                cell = sheet.cell(row=target_row, column=col_idx, value=numeric_value)
                cell.font = black_font  # --- APPLY FONT ---
            except (ValueError, TypeError):
                # Not a number, write as string
                cell = sheet.cell(row=target_row, column=col_idx, value=data_row.get(field_name, ""))
                cell.font = black_font  # --- APPLY FONT ---
        else:
            print(f"Warning: Header '{header_name}' not found in sheet '{sheet.title}'. Skipping data.")


def get_or_create_sheet(workbook, batea_name, template_sheet_name="TEMPLATE"):
    """
    Tries to find a sheet with batea_name. If not found,
    copies the internal 'TEMPLATE' sheet and renames it.
    """
    if batea_name in workbook.sheetnames:
        return workbook[batea_name]
    else:
        print(f"Sheet '{batea_name}' not found. Creating it from '{template_sheet_name}'...")
        
        if template_sheet_name not in workbook.sheetnames:
            print(f"CRITICAL ERROR: Template sheet '{template_sheet_name}' not found in the workbook.")
            print("Cannot create new sheets.")
            # We can't exit the program here, so we return None and let the main loop handle it
            return None 
            
        template_sheet = workbook[template_sheet_name]
        
        # --- FIX: Add an explicit check for a valid worksheet object ---
        if not hasattr(template_sheet, 'parent') or template_sheet.parent != workbook:
            print(f"CRITICAL ERROR: The template sheet '{template_sheet_name}' is not a valid worksheet or "
                  f"does not belong to this workbook.")
            print("This can sometimes happen if the file is corrupt.")
            return None
        # --- END FIX ---

        # --- FIX: Pass the object reference directly just to be safe ---
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = batea_name
        return new_sheet
    
def main():
    """Main execution function."""
    print("Starting PDF processing...")
    
    # We need to find the config file, whether running as .py or .exe
    config_path = get_resource_path('config.json')
    config = load_config(config_path)
    
    if config is None:
        input("Press Enter to exit.")
        return

    input_folder = "input_pdfs"
    if not setup_directories(input_folder):
        input("Press Enter to exit.")
        return

    all_data = []
    
    # Get field names for the Excel columns
    field_names = [field['name'] for field in config['extraction_fields']]
    
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print(f"No PDF files found in '{input_folder}'.")
        input("Press Enter to exit.")
        return

    print(f"Found {len(pdf_files)} PDF(s) to process...")

    for filename in pdf_files:
        pdf_path = os.path.join(input_folder, filename)
        print(f"Processing '{filename}'...")
        data = extract_data_from_pdf(pdf_path, config['extraction_fields'])
        
        if data:
            data["FECHA"] = data["FECHA"].split(" ")[0]
            data['Source File'] = filename  # Add source filename for reference
            all_data.append(data)

    if not all_data:
        print("No data was successfully extracted from any PDF.")
        input("Press Enter to exit.")
        return

    print("Data extraction complete. Generating Excel file...")
    
    # DEBUG: Print the full DataFrame
    # Create a pandas DataFrame
    # df = pd.DataFrame(all_data)
    # print("\nExtracted DataFrame:")
    # try:
    #     # Print without truncation
    #     with pd.option_context('display.max_rows', None, 'display.max_columns', None, 'display.width', 0):
    #         print(df.to_string(index=False))
    # except Exception:
    #     # Fallback
    #     print(df)
    
     # --- 2. Get Template and Output File Paths ---
    template_path = get_resource_path('PLANTILLA.xlsx')
    if not os.path.exists(template_path):
        print(f"Error: Template file 'PLANTILLA.xlsx' not found.")
        print("Please make sure it is in the same directory as the .exe")
        input("Press Enter to exit.")
        return

    output_filename = input("Enter the name for the output Excel file (e.g., 'datos.xlsx'): ")
    if not output_filename.lower().endswith('.xlsx'):
        output_filename += '.xlsx'

    # --- 3. Determine Mode: Create New or Update Existing ---
    
    if os.path.exists(output_filename):
        # --- MODE 2: UPDATE EXISTING FILE ---
        print(f"File '{output_filename}' exists. Will update it.")
        
        # --- NEW BACKUP LOGIC START ---
        try:
            backup_dir = ".old"
            os.makedirs(backup_dir, exist_ok=True)
            
            # Get timestamp
            now = datetime.datetime.now()
            timestamp = now.strftime("%Y-%m-%d_%H%M%S")
            
            # Create new filename
            base_name, extension = os.path.splitext(output_filename)
            backup_filename = f"{base_name}_{timestamp}{extension}"
            backup_path = os.path.join(backup_dir, backup_filename)
            
            # Copy the file
            shutil.copy(output_filename, backup_path)
            print(f"Created backup of '{output_filename}' at '{backup_path}'")
            
        except Exception as e:
            print(f"Warning: Could not create backup for '{output_filename}'. Error: {e}")
            # Continue execution even if backup fails.
        # --- NEW BACKUP LOGIC END ---
        
        try:
            workbook = load_workbook(output_filename)
            # --- FIX: We no longer need to load the external template ---
            # template_wb = load_workbook(template_path)
            # template_sheet = template_wb.active 
        except (InvalidFileException, FileNotFoundError):
            print(f"Error: Could not open '{output_filename}'. It might be corrupt or not a valid .xlsx file.")
            input("Press Enter to exit.")
            return
        except Exception as e:
            print(f"Error loading workbook: {e}")
            input("Press Enter to exit.")
            return
        
        # --- MODIFIED: Check for TEMPLATE sheet and create if missing ---
        template_sheet_name = "TEMPLATE"
        if template_sheet_name not in workbook.sheetnames:
            print(f"Warning: '{template_sheet_name}' sheet not found in '{output_filename}'.")
            print("Attempting to create it from 'PLANTILLA.xlsx'...")
            
            try:
                # Load the external template file
                if not os.path.exists(template_path):
                     print(f"CRITICAL ERROR: 'PLANTILLA.xlsx' not found at {template_path}. Cannot create 'TEMPLATE' sheet.")
                     input("Press Enter to exit.")
                     return

                template_wb = load_workbook(template_path)
                source_sheet = template_wb.active
                
                # Create the new TEMPLATE sheet in the destination workbook
                target_sheet = workbook.create_sheet(title=template_sheet_name)
                
                # Manually copy all properties
                copy_sheet_properties(source_sheet, target_sheet)
                
                template_wb.close()
                print(f"Successfully created 'TEMPLATE' sheet in '{output_filename}'.")
            
            except Exception as e:
                print(f"CRITICAL ERROR: Failed to create 'TEMPLATE' sheet from 'PLANTILLA.xlsx'.")
                print(f"Error details: {e}")
                input("Press Enter to exit.")
                return
        # --- END MODIFICATION ---
            
        for data_row in all_data:
            batea = data_row.get("BATEA", "").strip()
            if not batea:
                print(f"Warning: PDF '{data_row['Source File']}' has no BATEA. Skipping.")
                continue
            
            # --- FIX: Update function call ---
            sheet = get_or_create_sheet(workbook, batea, template_sheet_name)
            
            if sheet is None:
                print(f"Skipping PDF '{data_row['Source File']}' due to missing template sheet.")
                continue
                
            header_map = get_header_map(sheet, header_row=4)
            
            # Find next empty row (start from row 5)
            target_row = 5            
            print(f"Writing data from '{data_row['Source File']}' to sheet '{batea}', row {target_row}...")
            write_data_to_sheet(sheet, header_map, data_row, target_row)
        
        # template_wb.close() # No longer needed

    else:
        # --- MODE 1: CREATE NEW FILE ---
        print(f"Creating new file '{output_filename}' from template...")
        
        # Group data by BATEA
        data_by_batea = {}
        for row in all_data:
            batea = row.get('BATEA', 'UNKNOWN_BATEA').strip()
            if not batea:
                batea = 'UNKNOWN_BATEA'
            if batea not in data_by_batea:
                data_by_batea[batea] = []
            data_by_batea[batea].append(row)
        
        # Copy template to the new output filename
        shutil.copy(template_path, output_filename)
        
        try:
            # --- FIX START ---
            # We only load the new output file
            workbook = load_workbook(output_filename)
            
            # Get the first sheet (which is our template) and rename it.
            template_sheet = workbook.active
            template_sheet_name = "TEMPLATE"
            template_sheet.title = template_sheet_name
            # --- FIX END ---
            
        except (InvalidFileException, FileNotFoundError):
            print(f"Error: Could not open the new file '{output_filename}'.")
            input("Press Enter to exit.")
            return
        except Exception as e:
            print(f"Error loading new workbook: {e}")
            input("Press Enter to exit.")
            return

        # is_first_sheet = True # No longer needed
        for batea_name, batea_data in data_by_batea.items():
            # --- FIX: Always copy the template sheet ---
            print(f"Creating sheet '{batea_name}'...")
            sheet = get_or_create_sheet(workbook, batea_name, template_sheet_name)
            
            if sheet is None:
                print(f"Skipping Batea '{batea_name}' due to missing template sheet.")
                continue
            
            header_map = get_header_map(sheet, header_row=4)
            
            # Write all data for this batea, starting at row 5
            target_row = 5
            for data_row in batea_data:
                print(f"Writing data from '{data_row['Source File']}' to sheet '{batea_name}', row {target_row}...")
                write_data_to_sheet(sheet, header_map, data_row, target_row)
        
        # template_wb.close() # No longer needed

    # --- 4. Save and Exit ---
    try:
        # --- NEW: Remove the template sheet before saving ---
        if "TEMPLATE" in workbook.sheetnames:
            print("Removing internal 'TEMPLATE' sheet...")
            del workbook["TEMPLATE"]
        # --- END NEW ---

        workbook.save(output_filename)
        print(f"\nSuccess! Data saved to '{output_filename}'")
    except PermissionError:
        print(f"\nError: Could not save '{output_filename}'.")
        print("Please make sure the file is not open in Excel.")
    except Exception as e:
        print(f"\nError saving Excel file: {e}")

    workbook.close()
    print("Processing finished.")
    input("Press Enter to exit.")

if __name__ == "__main__":
    main()